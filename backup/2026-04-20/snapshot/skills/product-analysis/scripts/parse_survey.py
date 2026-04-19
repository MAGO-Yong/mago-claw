#!/usr/bin/env python3
"""
parse_survey.py - Parse and quality-score survey Excel files for product analysis.

Usage:
    python3 parse_survey.py <survey.xlsx> [--sheet 0] [--output report.json]

Output:
    JSON with:
    - headers: list of column names
    - records: list of dicts with all fields + computed scores
      - _fill_seconds: fill duration in seconds
      - _open_ended_count: number of open-ended questions answered
      - _open_ended_word_count: total words in open-ended answers
      - _quality_score: 0-100 composite score
      - _quality_tier: HIGH / MID / LOW / INVALID
      - _invalid_reason: reason string if INVALID, else null
"""

import sys
import json
import re
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def detect_open_ended_cols(headers):
    """Detect open-ended (free text) question column indices."""
    keywords = ["问答", "描述", "请回忆", "请说明", "其他", "填写", "补充", "原因", "场景"]
    exclude = ["多选", "单选", "评分"]
    indices = []
    for i, h in enumerate(headers):
        if h and any(kw in str(h) for kw in keywords) and not any(ex in str(h) for ex in exclude):
            indices.append(i)
    return indices


def detect_fill_time_col(headers):
    """Detect fill duration column index."""
    for i, h in enumerate(headers):
        if h and ("用时" in str(h) or "duration" in str(h).lower() or "seconds" in str(h).lower()):
            return i
    return None


def detect_name_col(headers):
    """Detect respondent name column index."""
    for i, h in enumerate(headers):
        if h and ("姓名" in str(h) or "name" in str(h).lower() or "员工" in str(h)):
            return i
    return None


def is_test_string(s):
    """Check if a string is clearly a test/placeholder value."""
    if not s:
        return False
    s = str(s).strip()
    test_patterns = [
        r'^[0-9]+$',           # pure numbers
        r'^[-—_]+$',           # dashes only
        r'^测试.*$',            # starts with 测试
        r'^test.*$',           # starts with test
        r'^[a-zA-Z]$',        # single letter
        r'^.$',                # single character
    ]
    return any(re.match(p, s, re.IGNORECASE) for p in test_patterns)


def score_answer_specificity(text):
    """Score answer specificity: 0=empty, 1=keyword, 2=scenario."""
    if not text or not str(text).strip():
        return 0
    text = str(text).strip()
    # Single word or very short → keyword level
    if len(text) < 10:
        return 1
    # Contains action verbs or scenario markers → scenario level
    scenario_markers = ["当", "查询", "排查", "发现", "需要", "希望", "流程", "问题", "自动", "操作"]
    if any(m in text for m in scenario_markers):
        return 2
    return 1


def compute_quality_score(fill_seconds, open_ended_count, word_count, specificity_avg,
                           max_seconds=3600, max_words=500, max_oe=10):
    """Compute 0-100 quality score."""
    # Cap extreme values (e.g., someone who left the form open for hours)
    capped_seconds = min(fill_seconds or 0, max_seconds)

    time_score = (capped_seconds / max_seconds) * 100 * 0.30
    word_score = min(word_count / max_words, 1.0) * 100 * 0.40
    oe_score = min(open_ended_count / max(max_oe, 1), 1.0) * 100 * 0.20
    spec_score = (specificity_avg / 2.0) * 100 * 0.10

    return round(time_score + word_score + oe_score + spec_score, 1)


def classify_invalid(fill_seconds, open_ended_answers):
    """Return (is_invalid, reason) tuple."""
    non_empty = [a for a in open_ended_answers if a and str(a).strip()]

    # All open-ended empty
    if not non_empty:
        return True, "所有问答题均为空"

    # All are test strings
    if all(is_test_string(a) for a in non_empty):
        return True, f"所有问答内容为测试/无效字符（{', '.join(str(a)[:10] for a in non_empty[:3])}）"

    # Extremely fast fill + near-empty content
    total_words = sum(len(str(a).split()) for a in non_empty)
    if fill_seconds and fill_seconds < 30 and total_words < 5:
        return True, f"填写时间过短（{fill_seconds}s）且内容极少（{total_words}词）"

    return False, None


def parse_survey(filepath, sheet_index=0):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[sheet_index]

    headers = [cell.value for cell in ws[1]]
    open_ended_cols = detect_open_ended_cols(headers)
    fill_time_col = detect_fill_time_col(headers)
    name_col = detect_name_col(headers)

    print(f"Detected {len(open_ended_cols)} open-ended columns: {[headers[i] for i in open_ended_cols]}", file=sys.stderr)
    print(f"Fill time column: {headers[fill_time_col] if fill_time_col is not None else 'not found'}", file=sys.stderr)

    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        record = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

        fill_seconds = None
        if fill_time_col is not None and row[fill_time_col] is not None:
            try:
                fill_seconds = float(row[fill_time_col])
            except (ValueError, TypeError):
                pass

        open_ended_answers = [row[i] for i in open_ended_cols if i < len(row)]
        non_empty_oe = [a for a in open_ended_answers if a and str(a).strip()]

        word_count = sum(len(str(a).split()) for a in non_empty_oe)
        specificities = [score_answer_specificity(a) for a in non_empty_oe]
        specificity_avg = sum(specificities) / len(specificities) if specificities else 0

        is_invalid, invalid_reason = classify_invalid(fill_seconds, open_ended_answers)

        quality_score = 0 if is_invalid else compute_quality_score(
            fill_seconds, len(non_empty_oe), word_count, specificity_avg
        )

        record["_fill_seconds"] = fill_seconds
        record["_open_ended_count"] = len(non_empty_oe)
        record["_open_ended_word_count"] = word_count
        record["_quality_score"] = quality_score
        record["_is_invalid"] = is_invalid
        record["_invalid_reason"] = invalid_reason
        record["_name"] = row[name_col] if name_col is not None else f"Row{row_idx}"

        records.append(record)

    # Tier assignment (after collecting all scores)
    valid_scores = sorted([r["_quality_score"] for r in records if not r["_is_invalid"]], reverse=True)
    high_threshold = valid_scores[int(len(valid_scores) * 0.25)] if valid_scores else 0
    mid_threshold = valid_scores[int(len(valid_scores) * 0.75)] if valid_scores else 0

    for r in records:
        if r["_is_invalid"]:
            r["_quality_tier"] = "INVALID"
        elif r["_quality_score"] >= high_threshold:
            r["_quality_tier"] = "HIGH"
        elif r["_quality_score"] >= mid_threshold:
            r["_quality_tier"] = "MID"
        else:
            r["_quality_tier"] = "LOW"

    return {
        "headers": headers,
        "open_ended_columns": [headers[i] for i in open_ended_cols],
        "fill_time_column": headers[fill_time_col] if fill_time_col is not None else None,
        "total_records": len(records),
        "invalid_count": sum(1 for r in records if r["_is_invalid"]),
        "tier_distribution": {
            "HIGH": sum(1 for r in records if r["_quality_tier"] == "HIGH"),
            "MID": sum(1 for r in records if r["_quality_tier"] == "MID"),
            "LOW": sum(1 for r in records if r["_quality_tier"] == "LOW"),
            "INVALID": sum(1 for r in records if r["_quality_tier"] == "INVALID"),
        },
        "records": records,
    }


def main():
    parser = argparse.ArgumentParser(description="Parse and quality-score survey Excel files.")
    parser.add_argument("filepath", help="Path to the Excel survey file")
    parser.add_argument("--sheet", type=int, default=0, help="Sheet index (0-based)")
    parser.add_argument("--output", help="Output JSON file path (default: stdout)")
    args = parser.parse_args()

    result = parse_survey(args.filepath, args.sheet)

    output = json.dumps(result, ensure_ascii=False, indent=2, default=str)
    if args.output:
        Path(args.output).write_text(output, encoding="utf-8")
        print(f"Written to {args.output}", file=sys.stderr)
    else:
        print(output)


if __name__ == "__main__":
    main()
